from openpyxl import Workbook
import subprocess

wb = Workbook()
ws = wb.create_sheet("gjnjkjr")

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

x = 1
words = []
info = []

while x <= 10:
    print("type your word №", x)
    words.append(input())
    x = x + 1

person = input("name yourself. - ")
subprocess.call(["git", "config", "--global", "user.name", person])

mail = input("type in your email. - ")
subprocess.call(["git", "config", "--global", "user.email", mail])

print(words, "words of", person,"and their email is :", mail)

ws.append(words)
info.append(person)
info.append(mail)
ws.append(info)

print("saving file...")
wb.save('DesCommitov.xlsx')
print("success, open the Excel file now.")