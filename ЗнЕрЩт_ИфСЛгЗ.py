from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet("Vtcctlp")
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
ws['A1'] = "Hello"
ws['B2'] = "everyone"
ws['C3'] = "in the"
ws['D4'] = "world"
ws['D1'] = input()
wb.save('rfrfirbyc.xlsx')
